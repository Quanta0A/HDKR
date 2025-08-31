from calendar import monthrange
from .utils.hdkr_calc import (
    calculate_io,
    calculate_hdkr,
    compute_daily_radiation,
    compute_monthly_radiation,
    erbs_diffuse_fraction,
)
from .utils.plotting import (
    plot_tilted_radiation,
    plot_radiation_vs_tilt,
    plot_hd_hb_it_bars,
    plot_optimal_tilt,
)
from .forms import RadiationForm
from django.shortcuts import render
from django.http import HttpResponse
import datetime
import math
import csv
import json
import logging
from io import BytesIO
from openpyxl import load_workbook
from django.contrib import messages

logger = logging.getLogger(__name__)

MONTHS = ['jan', 'feb', 'mar', 'apr', 'may', 'jun',
          'jul', 'aug', 'sep', 'oct', 'nov', 'dec']

# --------------------------------------------------------------------------- #
# HELPER: Parse uploaded CSV or Excel for full_month or 365_day modes        #
# --------------------------------------------------------------------------- #
def parse_uploaded_file(file, ghi_unit, expect_days):
    ghi_vals = []
    sun_vals = []
    file_name = file.name.lower()

    if file_name.endswith(".csv"):
        decoded_file = file.read().decode("utf-8").splitlines()
        reader = csv.DictReader(decoded_file)
        for row in reader:
            ghi_vals.append(float(row["GHI"]))
            if 'Sunshine' in row and row['Sunshine'].strip():
                sun_vals.append(float(row['Sunshine']))

    elif file_name.endswith(".xlsx"):
        wb = load_workbook(filename=BytesIO(file.read()), data_only=True)
        sheet = wb.active
        headers = [cell.value for cell in sheet[1]]
        ghi_idx = headers.index('GHI')
        sun_idx = headers.index('Sunshine') if 'Sunshine' in headers else None
        for row in sheet.iter_rows(min_row=2, values_only=True):
            ghi_vals.append(float(row[ghi_idx]))
            if sun_idx is not None and row[sun_idx] is not None:
                sun_vals.append(float(row[sun_idx]))

    else:
        raise ValueError("Unsupported file format. Please upload .csv or .xlsx")

    if ghi_unit == 'W':
        if len(ghi_vals) != expect_days or len(sun_vals) != expect_days:
            raise ValueError(f"Expected {expect_days} GHI and Sunshine values")
        ghi_vals = [(g * s * 3600) / 1e6 for g, s in zip(ghi_vals, sun_vals)]
    else:
        if len(ghi_vals) != expect_days:
            raise ValueError(f"Expected {expect_days} GHI values")

    return ghi_vals

# --------------------------------------------------------------------------- #
# MAIN VIEW                                                                   #
# --------------------------------------------------------------------------- #
def index(request):
    result = []
    graph = bar_graph = tilt_graph = optimal_tilt_graph = None

    if request.method == 'POST':
        form = RadiationForm(request.POST, request.FILES)
        if not form.is_valid():
            return render(request, 'solar_calc/index.html', {'form': form})

        lat   = form.cleaned_data['latitude']
        tilt  = form.cleaned_data['tilt']
        mode  = form.cleaned_data['mode']          # single_day / full_month / 12_month / 365_days
        ghi_unit = form.cleaned_data['ghi_unit']   # MJ or W
        ghi_raw  = form.cleaned_data['ghi']
        sun_raw  = form.cleaned_data['sunshine_hours']
        year     = form.cleaned_data['year'] or datetime.datetime.now().year
        albedo   = 0.2
        tilt_analysis = request.POST.get('tilt_analysis')

        if mode == '365_days':
            year_input_mode = 'daily'
        else:
            year_input_mode = request.POST.get('year_input_mode') or 'monthly'

        lat_rad  = math.radians(lat)
        tilt_rad = math.radians(tilt)

        if mode in ['12_month', '365_days']:
            mode = 'full_year'

        # ================================================================
        # FULL-YEAR BRANCH  (12-month or 365-day)
        # ================================================================
        if mode == 'full_year':
            try:
                if year_input_mode == 'monthly':
                    ghi_vals_mj = []
                    for m in MONTHS:
                        g_raw = request.POST.get(f'month_{m}_ghi')
                        s_raw = request.POST.get(f'month_{m}_sunshine')
                        if not g_raw:
                            form.add_error(None, f'Missing GHI for {m.capitalize()}')
                            return render(request, 'solar_calc/index.html', {'form': form})

                        if ghi_unit == 'W':
                            if not s_raw:
                                form.add_error(None, f'Missing sunshine for {m.capitalize()}')
                                return render(request, 'solar_calc/index.html', {'form': form})
                            g = float(g_raw)
                            s = float(s_raw)
                            ghi_vals_mj.append((g * s * 3600) / 1e6)
                        else:
                            ghi_vals_mj.append(float(g_raw))

                    results = compute_monthly_radiation(ghi_vals_mj, lat, tilt, albedo)
                    label   = '12-Month Average'
                    optimal_tilt_graph = plot_optimal_tilt(results, mode='monthly')

                elif year_input_mode == 'daily':
                    try:
                        csv_file = request.FILES.get('csv_file')
                        if csv_file:
                            ghi_vals = parse_uploaded_file(csv_file, ghi_unit, 365)
                        else:
                            ghi_vals = [float(x.strip()) for x in ghi_raw.split(',') if x.strip()]
                            if ghi_unit == 'W':
                                sun_vals = [float(x.strip()) for x in sun_raw.split(',') if x.strip()]
                                if len(ghi_vals) != 365 or len(sun_vals) != 365:
                                    raise ValueError("365 GHI and Sunshine values required")
                                ghi_vals = [(g * s * 3600) / 1e6 for g, s in zip(ghi_vals, sun_vals)]
                            elif len(ghi_vals) != 365:
                                raise ValueError("365 GHI values required")

                        results = compute_daily_radiation(ghi_vals, lat, tilt, albedo, start_day=1)
                        label   = 'Full Year (365 Days)'

                        if tilt_analysis:
                            tilt_results = []
                            for tilt_angle in range(0, 91):
                                t_rad = math.radians(tilt_angle)

                                # FIXED: separate lists
                                Hd_list, Hb_list, It_list = [], [], []

                                for day in results:
                                    Hd_day = day['Hd']
                                    Hb_day = day['Hb']
                                    H_day = Hd_day + Hb_day
                                    delta_rad = math.radians(day['declination'])
                                    vals = calculate_hdkr(H_day, Hd_day, lat_rad, t_rad, delta_rad, albedo)
                                    Hd_list.append(vals['Hd_tilted'])
                                    Hb_list.append(vals['Hb_tilted'])
                                    It_list.append(vals['It'])

                                n = len(results)
                                tilt_results.append({
                                    'tilt': tilt_angle,
                                    'Hd': sum(Hd_list) / n,
                                    'Hb': sum(Hb_list) / n,
                                    'It': sum(It_list) / n
                                })

                            tilt_graph = plot_radiation_vs_tilt(tilt_results)
                            optimal_tilt_graph = plot_optimal_tilt(tilt_results)

                    except Exception as e:
                        form.add_error('csv_file', f"Yearly data error: {e}")
                        return render(request, 'solar_calc/index.html', {'form': form})

                else:
                    form.add_error(None, 'Invalid yearly input mode.')
                    return render(request, 'solar_calc/index.html', {'form': form})

                result = results
                graph  = plot_tilted_radiation(results, label=label)
                bar_graph = plot_hd_hb_it_bars(results, label=label)

            except Exception as e:
                form.add_error(None, f'Yearly processing error: {e}')
                return render(request, 'solar_calc/index.html', {'form': form})

        # ================================================================
        # FULL-MONTH
        # ================================================================
        elif mode == 'full_month':
            month = int(form.cleaned_data['month'])
            num_days = monthrange(year, month)[1]
            ghi_vals = []
            sun_vals = []

            csv_file = request.FILES.get('csv_file')
            if csv_file:
                try:
                    ghi_vals = parse_uploaded_file(csv_file, ghi_unit, num_days)
                    messages.success(request, f"{csv_file.name} uploaded successfully.")
                except Exception as e:
                    form.add_error('csv_file', f"Full-month data error: {e}")
                    return render(request, 'solar_calc/index.html', {'form': form})
            else:
                try:
                    ghi_vals = [float(g.strip()) for g in ghi_raw.split(',') if g.strip()]
                    if ghi_unit == 'W':
                        sun_vals = [float(s.strip()) for s in sun_raw.split(',') if s.strip()]
                        if len(ghi_vals) != len(sun_vals):
                            form.add_error('ghi', 'GHI and sunshine count must match.')
                            return render(request, 'solar_calc/index.html', {'form': form})
                        ghi_vals = [(g * s * 3600) / 1e6 for g, s in zip(ghi_vals, sun_vals)]
                except ValueError:
                    form.add_error('ghi', 'Invalid GHI or Sunshine input format.')
                    return render(request, 'solar_calc/index.html', {'form': form})

            if len(ghi_vals) != num_days:
                form.add_error('ghi', f'Enter exactly {num_days} GHI values.')
                return render(request, 'solar_calc/index.html', {'form': form})

            result = []
            for day in range(1, num_days + 1):
                H = ghi_vals[day - 1]
                day_num = datetime.date(year, month, day).timetuple().tm_yday
                io, delta, delta_rad = calculate_io(day_num, lat)
                kt = H / io if io else 0
                Hd_H = erbs_diffuse_fraction(kt)
                Hd = Hd_H * H
                values = calculate_hdkr(H, Hd, lat_rad, tilt_rad, delta_rad, albedo)

                result.append({
                    'day': f"{day:02d}-{month:02d}",
                    'declination': delta,
                    'Io': io,
                    'Kt': kt,
                    'Hd_H': Hd_H,
                    'Hd': values['Hd'],
                    'Hb': values['Hb'],
                    'rb': values['rb'],
                    'Hd_tilted': values['Hd_tilted'],
                    'Hb_tilted': values['Hb_tilted'],
                    'It': values['It']
                })

            label = datetime.date(year, month, 1).strftime('%B %Y')
            graph = plot_tilted_radiation(result, label=label)
            bar_graph = plot_hd_hb_it_bars(result, label=label)

            if tilt_analysis:
                tilt_results = []
                for tilt_angle in range(0, 91):
                    tilt_rad_tmp = math.radians(tilt_angle)

                    # FIXED: separate lists
                    Hd_list, Hb_list, It_list = [], [], []

                    for day in range(1, num_days + 1):
                        H = ghi_vals[day - 1]
                        day_num = datetime.date(year, month, day).timetuple().tm_yday
                        io, delta, delta_rad = calculate_io(day_num, lat)
                        kt = H / io if io else 0
                        Hd_H = erbs_diffuse_fraction(kt)
                        Hd = Hd_H * H
                        vals = calculate_hdkr(H, Hd, lat_rad, tilt_rad_tmp, delta_rad, albedo)
                        Hd_list.append(vals['Hd_tilted'])
                        Hb_list.append(vals['Hb_tilted'])
                        It_list.append(vals['It'])

                    tilt_results.append({
                        'tilt': tilt_angle,
                        'Hd': sum(Hd_list) / num_days,
                        'Hb': sum(Hb_list) / num_days,
                        'It': sum(It_list) / num_days
                    })

                tilt_graph = plot_radiation_vs_tilt(tilt_results)
                optimal_tilt_graph = plot_optimal_tilt(tilt_results)

        # ================================================================
        # SINGLE-DAY MODE
        # ================================================================
        elif mode == 'single_day':
            date = form.cleaned_data['date']
            if not date:
                form.add_error('date', 'Please select a valid date.')
                return render(request, 'solar_calc/index.html', {'form': form})

            day_of_year = date.timetuple().tm_yday
            try:
                if ghi_unit == 'MJ':
                    H = float(ghi_raw.strip())
                else:
                    g = float(ghi_raw.strip())
                    s = float(sun_raw.strip())
                    H = (g * s * 3600) / 1e6
            except ValueError:
                form.add_error('ghi', 'Invalid GHI input.')
                return render(request, 'solar_calc/index.html', {'form': form})

            io, delta, delta_rad = calculate_io(day_of_year, lat)
            kt = H / io if io else 0
            Hd_H = erbs_diffuse_fraction(kt)
            Hd = Hd_H * H
            values = calculate_hdkr(H, Hd, lat_rad, tilt_rad, delta_rad, albedo)

            result = [{
                'day': date.strftime('%d-%b'),
                'declination': delta,
                'Io': io,
                'Kt': kt,
                'Hd_H': Hd_H,
                'Hd': values['Hd'],
                'Hb': values['Hb'],
                'rb': values['rb'],
                'Hd_tilted': values['Hd_tilted'],
                'Hb_tilted': values['Hb_tilted'],
                'It': values['It']
            }]
            label = date.strftime('%d %B %Y')
            graph = plot_tilted_radiation(result, label=label)

            if tilt_analysis:
                tilt_results = []
                for tilt_angle in range(0, 91):
                    t_rad = math.radians(tilt_angle)
                    vals = calculate_hdkr(H, Hd, lat_rad, t_rad, delta_rad, albedo)

                    # Store individual tilted values
                    tilt_results.append({
                        'tilt': tilt_angle,
                        'Hd': vals['Hd_tilted'],
                        'Hb': vals['Hb_tilted'],
                        'It': vals['It']
                    })
                tilt_graph = plot_radiation_vs_tilt(tilt_results)
                optimal_tilt_graph = plot_optimal_tilt(tilt_results)

    else:
        form = RadiationForm()

    return render(request, 'solar_calc/index.html', {
        'form': form,
        'result': result,
        'graph': graph,
        'bar_graph': bar_graph,
        'tilt_graph': tilt_graph,
        'optimal_tilt_graph': optimal_tilt_graph,
        'months': MONTHS,
    })

# --------------------------------------------------------------------------- #
# CSV DOWNLOAD VIEW                                                           #
# --------------------------------------------------------------------------- #
def download_csv(request):
    if request.method == 'POST':
        result_json = request.POST.get('result_json')
        if not result_json:
            return HttpResponse('No data to download', status=400)

        try:
            results = json.loads(result_json)
        except json.JSONDecodeError:
            return HttpResponse('Invalid JSON data', status=400)

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="solar_radiation_results.csv"'

        writer = csv.DictWriter(response, fieldnames=list(results[0].keys()))
        writer.writeheader()
        writer.writerows(results)
        return response

    return HttpResponse('Invalid request', status=405)
